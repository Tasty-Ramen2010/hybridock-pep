"""E18 v2 — orchestrator: Stage 1+ (bond-strength SASA) + Stage 2+ (MD entropy).

Per complex computes, on the actual 3D structure:
  Stage 1+  de_sasa_strength : Σ_i ΔSASA_i · favorability_i   (favorable burial -,
            unfavorable buried-unsat-polar / like-charge +)   + clash_pen
  Stage 2+  dS_dih  = Σ_i (S_bound_i − S_free_i)   (dihedral-histogram entropy, MD)
            rmsf_ratio = mean(RMSF_free)/mean(RMSF_bound)  (rigidification on binding)
Carries y, L, grp, seq + geometry (hb_count, aromatic_cc) for the ablation baseline.

Incrementally checkpoints to /tmp/e18v2_{cr,pb}.json so a crash loses nothing.
Runtime ~40-50 s/complex (bound MD + free MD). Scope: crystal-65 + resolvable PEPBI groups.
"""
from __future__ import annotations

import glob
import json
import os
import sys
import time
import warnings
from pathlib import Path

import numpy as np

warnings.filterwarnings("ignore")
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "scripts"))
from Bio.PDB import NeighborSearch, PDBIO, PDBParser, Select  # noqa: E402
from Bio.PDB.SASA import ShrakeRupley  # noqa: E402
from e18v2_md import run_bound_dynamics, run_free_dynamics  # noqa: E402
from e18_hybrid_features import EISENBERG, AA3to1  # noqa: E402

P = PDBParser(QUIET=True)
SR = ShrakeRupley()
IO = PDBIO()
POS = {"ARG", "LYS", "HIS"}
NEG = {"ASP", "GLU"}
CHARGED = POS | NEG
POLAR = {"SER", "THR", "ASN", "GLN", "TYR", "CYS", "TRP", "HIS"}
APOLAR = {"ALA", "VAL", "LEU", "ILE", "MET", "PHE", "PRO", "GLY", "TRP"}
AROM = {"PHE", "TYR", "TRP", "HIS"}


def _per_res_sasa(struct):
    SR.compute(struct, level="A")
    return {(r.get_parent().id, r.id): sum(float(a.sasa) for a in r)
            for r in struct.get_residues() if r.id[0] == " "}


def bond_strength_sasa(pep_pdb, complex_path, pep_chain):
    """Stage 1+: per-residue ΔSASA weighted by contact favorability + clash penalty."""
    free = _per_res_sasa(P.get_structure("f", pep_pdb))
    cpx = _per_res_sasa(P.get_structure("c", complex_path))
    cx = P.get_structure("cc", complex_path)[0]
    pep_res = [r for r in cx[pep_chain] if r.id[0] == " "]
    rec_atoms = [a for ch in cx if ch.id != pep_chain for r in ch if r.id[0] == " "
                 for a in r if a.element != "H"]
    if not pep_res or not rec_atoms:
        return None
    ns = NeighborSearch(rec_atoms)
    pf = [r for r in P.get_structure("pp", pep_pdb)[0].get_residues() if r.id[0] == " "]
    n = min(len(pep_res), len(pf))
    de = 0.0
    clash = 0
    for i in range(n):
        rc = pep_res[i]
        rn = rc.resname.upper()
        aa = AA3to1.get(rn, "A")
        rfree = free.get((pf[i].get_parent().id, pf[i].id), 0.0)
        rbound = cpx.get((rc.get_parent().id, rc.id), 0.0)
        dsasa = max(0.0, rfree - rbound)  # buried area for residue i
        if dsasa < 1.0:
            continue
        # contacts for this residue
        has_hb = has_sb = has_apolar = False
        unsat_polar = (rn in POLAR or rn in CHARGED)
        for a in rc:
            if a.element == "H":
                continue
            near = ns.search(a.coord, 4.5)
            for b in near:
                d = np.linalg.norm(a.coord - b.coord)
                if d < 2.5:
                    clash += 1
                brn = b.get_parent().resname.upper()
                if a.element in ("N", "O") and b.element in ("N", "O") and d <= 3.5:
                    has_hb = True
                    unsat_polar = False
                if rn in CHARGED and a.element in ("N", "O") and d <= 4.0:
                    want = NEG if rn in POS else POS
                    if brn in want:
                        has_sb = True
                        unsat_polar = False
                if rn in APOLAR and brn in APOLAR and d <= 5.0:
                    has_apolar = True
        # favorability multiplier
        base = -EISENBERG.get(aa, 0.0)          # hydrophobic burial favorable (neg)
        fav = base
        if has_hb:
            fav -= 0.3                            # satisfied H-bond: extra favorable
        if has_sb:
            fav -= 0.6                            # salt bridge: strongly favorable
        if has_apolar:
            fav -= 0.2                            # hydrophobic packing
        if unsat_polar:
            fav += 1.0                            # buried unsatisfied polar/charge: PENALTY
        de += fav * dsasa
    de /= 100.0
    return dict(de_strength=de, clash_pen=float(clash))


# ---------------- dataset processing ----------------

class _ChainB(Select):
    def accept_chain(self, ch): return ch.id == "B"
    def accept_residue(self, res): return res.id[0] == " "


def _geom_hb_arom(pep_res, rec_atoms, hb_cut=3.5):
    if not pep_res or not rec_atoms:
        return 0, 0
    ns = NeighborSearch(rec_atoms)
    hb = arom = 0
    for rp in pep_res:
        for a in rp:
            if a.element in ("N", "O") and any(
                    b.element in ("N", "O") and np.linalg.norm(a.coord - b.coord) <= hb_cut
                    for b in ns.search(a.coord, hb_cut)):
                hb += 1
        if rp.resname.upper() in AROM and any(
                b.get_parent().resname.upper() in AROM
                for a in rp if a.element != "H" for b in ns.search(a.coord, 5.5)):
            arom += 1
    return hb, arom


def crystal_records(prod_ps):
    e0 = json.loads(Path("/tmp/e0_rows.json").read_text())
    base = json.loads((ROOT / "data/benchmark_crystal.json").read_text())
    sm = {b["pdb"].upper(): b["peptide_seq"] for b in base}
    e14 = {r["seq"]: r for r in json.loads(Path("/tmp/e14_cr.json").read_text())}
    out_path = Path("/tmp/e18v2_cr.json")
    out = json.loads(out_path.read_text()) if out_path.exists() else []
    done = {r["pdb"] for r in out}
    for r in e0:
        if not r.get("pep_pdb") or r["pdb"].upper() in done:
            continue
        pdb = r["pdb"].upper()
        seq = sm.get(pdb, "")
        if not seq:
            continue
        merged = Path(f"/tmp/e18v2_cx/{pdb}.pdb"); merged.parent.mkdir(exist_ok=True)
        lines = []
        for src, ch in ((r["pep_pdb"], "P"), (r["poc_pdb"], "R")):
            for ln in Path(src).read_text().splitlines():
                if ln.startswith(("ATOM", "HETATM")) and ln[17:20] != "HOH":
                    lines.append(ln[:21] + ch + ln[22:])
        merged.write_text("\n".join(lines) + "\nEND\n")
        try:
            s1 = bond_strength_sasa(r["pep_pdb"], str(merged), "P")
            if s1 is None:
                continue
            rb, sb = run_bound_dynamics(r["pep_pdb"], r["poc_pdb"], prod_ps)
            rf, sf = run_free_dynamics(r["pep_pdb"], prod_ps)
            ds_dih = float(np.nansum(sb - sf))
            rmsf_ratio = float(np.nanmean(rf) / (np.nanmean(rb) + 1e-6))
            g = e14.get(seq, {})
            rec = dict(pdb=pdb, y=r["y"], L=r["L"], aff=r["aff"], seq=seq,
                       grp=g.get("grp", f"cr_{pdb}"), hb_count=g.get("hb_count"),
                       aromatic_cc=g.get("aromatic_cc"), ds_dih=ds_dih,
                       rmsf_ratio=rmsf_ratio, **s1)
            out.append(rec)
            out_path.write_text(json.dumps(out))
            print(f"  cr {pdb}: de_str={s1['de_strength']:+.1f} dS_dih={ds_dih:+.2f} "
                  f"rmsf_ratio={rmsf_ratio:.2f} (n={len(out)})", flush=True)
        except Exception as e:  # noqa: BLE001
            print(f"  cr {pdb} FAIL {type(e).__name__}: {str(e)[:60]}", flush=True)
    return out


def resolvable_pepbi_groups():
    """Binding groups with >=4 members and ΔG spread > 1.5 kcal/mol."""
    import openpyxl
    wb = openpyxl.load_workbook("/tmp/pepbi/PEPBI.xlsx", read_only=True)
    rows = list(wb["PEPBI Data"].iter_rows(values_only=True))
    hdr = rows[1]; ci = lambda n: hdr.index(n)
    def num(x):
        try: return float(x)
        except (TypeError, ValueError): return None
    c_bg, c_dg, c_kd = ci("Binding Group"), ci("ΔG (kcal/mol)"), ci("KD (M)")
    bg = {}
    for r in rows[2:]:
        dg, kd = num(r[c_dg]), num(r[c_kd])
        if dg is None and kd and kd > 0:
            dg = 0.593 * np.log(kd)
        if dg is None:
            continue
        bg.setdefault(str(r[c_bg]), []).append(dg)
    return {g for g, dgs in bg.items() if len(dgs) >= 4 and (max(dgs) - min(dgs)) > 1.5}


def pepbi_records(prod_ps, max_per_group=10):
    files = {os.path.basename(f)[:-4].lower(): f
             for f in glob.glob("/tmp/pepbi/struct/**/*.pdb", recursive=True)}
    keep_groups = resolvable_pepbi_groups()
    import openpyxl
    wb = openpyxl.load_workbook("/tmp/pepbi/PEPBI.xlsx", read_only=True)
    rows = list(wb["PEPBI Data"].iter_rows(values_only=True))
    hdr = rows[1]; ci = lambda n: hdr.index(n)
    def num(x):
        try: return float(x)
        except (TypeError, ValueError): return None
    c_nm, c_dg, c_kd, c_bg, c_seq = (ci("PEPBI Complex Name"), ci("ΔG (kcal/mol)"),
                                     ci("KD (M)"), ci("Binding Group"), ci("Peptide Sequence"))
    # collect per group, spread across ΔG, cap
    by_group = {}
    for r in rows[2:]:
        bgv = str(r[c_bg])
        if bgv not in keep_groups:
            continue
        nm = str(r[c_nm]).strip().lower() if r[c_nm] else None
        if not nm or nm not in files:
            continue
        dg, kd = num(r[c_dg]), num(r[c_kd])
        if dg is None and kd and kd > 0:
            dg = 0.593 * np.log(kd)
        if dg is None:
            continue
        seq = str(r[c_seq]).strip().upper() if r[c_seq] else ""
        by_group.setdefault(bgv, []).append((nm, files[nm], dg, seq))
    out_path = Path("/tmp/e18v2_pb.json")
    out = json.loads(out_path.read_text()) if out_path.exists() else []
    done = {r["nm"] for r in out}
    pep_d = Path("/tmp/e18v2_pep"); pep_d.mkdir(exist_ok=True)
    for bgv, variants in by_group.items():
        variants = sorted(variants, key=lambda t: t[2])
        if len(variants) > max_per_group:
            idx = np.linspace(0, len(variants)-1, max_per_group).astype(int)
            variants = [variants[i] for i in idx]
        for nm, cx, dg, seq in variants:
            if nm in done:
                continue
            s = P.get_structure("x", cx)[0]
            if "B" not in [c.id for c in s]:
                continue
            io = PDBIO(); io.set_structure(P.get_structure("y", cx))
            pep_pdb = pep_d / f"{nm}.pdb"; io.save(str(pep_pdb), _ChainB())
            if not seq:
                seq = "".join(AA3to1.get(x.resname.upper(), "A") for x in s["B"] if x.id[0] == " ")
            # pocket-only file for MD (crop to 12Å like e17)
            poc_pdb = pep_d / f"{nm}_poc.pdb"
            _write_pocket(cx, str(poc_pdb))
            try:
                s1 = bond_strength_sasa(str(pep_pdb), cx, "B")
                if s1 is None:
                    continue
                # MD wants chain "P" peptide + "R" receptor; reuse bound dynamics with pep+poc
                rb, sb = run_bound_dynamics(str(pep_pdb), str(poc_pdb), prod_ps)
                rf, sf = run_free_dynamics(str(pep_pdb), prod_ps)
                ds_dih = float(np.nansum(sb - sf))
                rmsf_ratio = float(np.nanmean(rf) / (np.nanmean(rb) + 1e-6))
                pep_res = [x for x in s["B"] if x.id[0] == " "]
                rec_atoms = [a for c in s if c.id != "B" for x in c if x.id[0] == " "
                             for a in x if a.element != "H"]
                hb, arom = _geom_hb_arom(pep_res, rec_atoms)
                rec = dict(nm=nm, y=dg, L=len(seq), aff="Kd", seq=seq, grp=f"pb_{bgv}",
                           hb_count=hb, aromatic_cc=arom, ds_dih=ds_dih,
                           rmsf_ratio=rmsf_ratio, **s1)
                out.append(rec); out_path.write_text(json.dumps(out))
                print(f"  pb {nm}: de_str={s1['de_strength']:+.1f} dS_dih={ds_dih:+.2f} "
                      f"rmsf_ratio={rmsf_ratio:.2f} (n={len(out)})", flush=True)
            except Exception as e:  # noqa: BLE001
                print(f"  pb {nm} FAIL {type(e).__name__}: {str(e)[:60]}", flush=True)
    return out


def _write_pocket(cx, out, radius=12.0):
    s = P.get_structure("x", cx)
    pep_xyz = np.array([a.coord for r in s[0]["B"] if r.id[0] == " "
                        for a in r if a.element != "H"])
    keep = set()
    r2 = radius * radius
    for ch in s[0]:
        if ch.id == "B":
            continue
        for res in ch:
            if res.id[0] != " ":
                continue
            for a in res:
                if a.element != "H" and np.min(((pep_xyz - a.coord) ** 2).sum(1)) <= r2:
                    keep.add((ch.id, res.id)); break

    class _Poc(Select):
        def accept_chain(self, c): return c.id != "B"
        def accept_residue(self, res): return res.id[0] == " " and (res.get_parent().id, res.id) in keep
    io = PDBIO(); io.set_structure(s); io.save(out, _Poc())


def main():
    prod_ps = int(sys.argv[1]) if len(sys.argv) > 1 else 100
    which = sys.argv[2] if len(sys.argv) > 2 else "both"
    t0 = time.time()
    if which in ("both", "cr"):
        print("=== crystal-65 ===", flush=True)
        crystal_records(prod_ps)
    if which in ("both", "pb"):
        print("=== PEPBI resolvable groups ===", flush=True)
        pepbi_records(prod_ps)
    print(f"total {(time.time()-t0)/60:.1f} min", flush=True)


if __name__ == "__main__":
    main()
