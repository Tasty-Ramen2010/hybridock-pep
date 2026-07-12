"""E186 — ingest PPIKB (docs/Affinity Dataset(branch).xlsx) into a clean jsonl for training + selectivity.

Filters: Linear only (skip cyclic/chem-mod), standard-AA peptide seq, parse Affinity to ΔG (kcal/mol).
Keeps Protein_Sequence (for receptor/family grouping) + PDB_ID. Dedup. Writes data/ppikb_clean.jsonl.
ΔG = RT ln(K) with K in molar; T=298.15 → ΔG(kcal/mol) = 0.5925 * ln(value_nM * 1e-9).
"""
from __future__ import annotations

import json
import math
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = next(ROOT.glob("docs/*ffinity*.xlsx"))
RT = 0.5925  # kcal/mol at 298.15K
STD = set("ACDEFGHIKLMNPQRSTVWY")


def parse_aff(s):
    """'IC50=0.6 nM' / 'Kd=1000 nM' -> (type, dG_kcal). Returns None if unparseable."""
    if not s:
        return None
    m = re.match(r"\s*([A-Za-z0-9]+)\s*=\s*([\d.eE+-]+)\s*([numpfM]+)", str(s))
    if not m:
        return None
    typ, val, unit = m.group(1), float(m.group(2)), m.group(3)
    if val <= 0:
        return None
    scale = {"M": 1, "mM": 1e-3, "uM": 1e-6, "nM": 1e-9, "pM": 1e-12, "fM": 1e-15}.get(unit)
    if scale is None:
        return None
    molar = val * scale
    return typ, RT * math.log(molar)  # negative ΔG


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True); ws = wb["ppi_research"]
    H = {c.value: i for i, c in enumerate(next(ws.iter_rows(max_row=1)))}

    def g(r, k):
        return r[H[k]]

    seen = set(); out = []
    n_total = n_cyclic = n_badseq = n_badaff = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        n_total += 1
        if g(r, "Linear/Cyclic") != "Linear":
            n_cyclic += 1; continue
        seq = str(g(r, "Peptide_Sequence") or "").strip().upper()
        if not seq or any(c not in STD for c in seq) or not (2 <= len(seq) <= 50):
            n_badseq += 1; continue
        aff = parse_aff(g(r, "Affinity"))
        if aff is None:
            n_badaff += 1; continue
        typ, dg = aff
        pdb = str(g(r, "PDB_ID") or "").strip().lower()
        prot = str(g(r, "Protein_Sequence") or "").strip().upper()
        key = (pdb, seq, round(dg, 2))
        if key in seen:
            continue
        seen.add(key)
        out.append({"id": g(r, "ID"), "pdb": pdb, "seq": seq, "length": len(seq),
                    "y": round(dg, 3), "aff_type": typ, "protein_seq": prot,
                    "protein_name": str(g(r, "Protein_Name") or "")[:60],
                    "net_charge": sum(c in "KR" for c in seq) - sum(c in "DE" for c in seq)})
    with open(ROOT / "data/ppikb_clean.jsonl", "w") as fh:
        for o in out:
            fh.write(json.dumps(o) + "\n")

    print(f"PPIKB: {n_total} rows -> {len(out)} clean linear-AA-affinity entries")
    print(f"  skipped: cyclic/nonlinear={n_cyclic}, bad seq={n_badseq}, bad affinity={n_badaff}")
    from collections import Counter
    print(f"  aff types: {Counter(o['aff_type'] for o in out).most_common()}")
    # overlap with our 925
    ours = {json.loads(l)["pdb"].lower() for l in open(ROOT / "data/pdbbind_peptides.jsonl")}
    pdbs = {o["pdb"] for o in out if o["pdb"]}
    print(f"  unique PDBs: {len(pdbs)}, NOT in our 925: {len(pdbs - ours)}")
    # selectivity families: group by protein_seq
    fam = {}
    for o in out:
        fam.setdefault(o["protein_seq"][:50], []).append(o)
    multi = {k: v for k, v in fam.items() if len({x['seq'] for x in v}) >= 4}
    spread = {k: (max(x['y'] for x in v) - min(x['y'] for x in v)) for k, v in multi.items()}
    wide = {k: v for k, v in multi.items() if spread[k] >= 2.0}
    print(f"  selectivity families (>=4 distinct peptides): {len(multi)}; with >=2 kcal/mol spread: {len(wide)}")
    print(f"  total peptides in wide families: {sum(len(v) for v in wide.values())}")
    # Kd-only
    kd = [o for o in out if o["aff_type"] in ("Kd", "KD", "pKd")]
    print(f"  Kd-only entries: {len(kd)} ({len({o['pdb'] for o in kd})} PDBs)")


if __name__ == "__main__":
    main()
