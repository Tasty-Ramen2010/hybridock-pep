"""E192 — ingest the FULL PPIKB main dataset (19.5k rows, downloaded from ppikb.duanlab.ac) + two tests:
  (A) does SCALE unlock sequence selectivity? 454 families / 10250 peptides (vs 80/1100 branch, tau 0.059)
  (B) PPI-clone on the PPIKB structured subset — does PPI's own feature class do BETTER on PPIKB than us,
      or is PPIKB just a harder/noisier dataset for everyone (explains "why we get worse on new dataset")?
Writes data/ppikb_main_clean.jsonl (sequence-level, gitignored if >1MB).
"""
from __future__ import annotations

import importlib.util
import json
import math
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

for _v in ("OMP_NUM_THREADS", "OPENBLAS_NUM_THREADS", "MKL_NUM_THREADS"):
    os.environ[_v] = "3"
import numpy as np  # noqa: E402
import openpyxl  # noqa: E402
from scipy.stats import spearmanr  # noqa: E402
from sklearn.ensemble import HistGradientBoostingRegressor  # noqa: E402
from sklearn.model_selection import LeaveOneGroupOut  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))
e150 = importlib.util.module_from_spec(importlib.util.spec_from_file_location("e150", ROOT / "experiments/e150_protdcal_descriptors.py"))
importlib.util.spec_from_file_location("e150", ROOT / "experiments/e150_protdcal_descriptors.py").loader.exec_module(e150)
SD, POS, NEG = e150.seq_descriptors, e150.POS, e150.NEG
RT = 0.5925
STD = set("ACDEFGHIKLMNPQRSTVWY")


def parse(s):
    m = re.match(r"\s*([A-Za-z0-9]+)\s*=\s*([\d.eE+-]+)\s*([numpfM]+)", str(s))
    if not m:
        return None
    typ, val, unit = m.group(1), float(m.group(2)), m.group(3)
    sc = {"M": 1, "mM": 1e-3, "uM": 1e-6, "nM": 1e-9, "pM": 1e-12, "fM": 1e-15}.get(unit)
    if sc is None or val <= 0:
        return None
    return typ, RT * math.log(val * sc)


def ingest():
    wb = openpyxl.load_workbook(ROOT / "docs/Affinity Dataset(main).xlsx", read_only=True)
    ws = wb["ppi_research"]
    H = {c.value: i for i, c in enumerate(next(ws.iter_rows(max_row=1)))}
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[H["Linear/Cyclic"]] != "Linear":
            continue
        seq = str(r[H["Peptide_Sequence"]] or "").strip().upper()
        if not seq or any(c not in STD for c in seq) or not (2 <= len(seq) <= 50):
            continue
        a = parse(r[H["Affinity"]])
        if a is None:
            continue
        out.append({"seq": seq, "y": round(a[1], 3), "aff_type": a[0], "length": len(seq),
                    "protein_seq": str(r[H["Protein_Sequence"]] or "")[:60],
                    "net_charge": sum(c in "KR" for c in seq) - sum(c in "DE" for c in seq)})
    with open(ROOT / "data/ppikb_main_clean.jsonl", "w") as fh:
        for o in out:
            fh.write(json.dumps(o) + "\n")
    return out


def feat(seq):
    pq = sum(c in POS for c in seq) - sum(c in NEG for c in seq)
    return SD(seq) + [float(pq), float(abs(pq)), float(len(seq))]


def met(p, y):
    p, y = np.asarray(p, float), np.asarray(y, float)
    ok = ~(np.isnan(p) | np.isnan(y))
    return float(np.corrcoef(p[ok], y[ok])[0, 1]), float(np.mean(np.abs(p[ok] - y[ok])))


def main():
    cache = ROOT / "data/ppikb_main_clean.jsonl"
    rows = [json.loads(l) for l in open(cache)] if cache.exists() else ingest()
    print(f"PPIKB-main clean: {len(rows)} entries\n")

    # (A) scale selectivity: families >=4 peptides, >=2 kcal spread
    fam = defaultdict(list)
    for r in rows:
        fam[r["protein_seq"][:50]].append(r)
    fams = [(k, v) for k, v in fam.items() if len({x["seq"] for x in v}) >= 4
            and (max(x["y"] for x in v) - min(x["y"] for x in v)) >= 2.0]
    X = np.nan_to_num([feat(r["seq"]) for _, v in fams for r in v])
    y = np.array([r["y"] for _, v in fams for r in v])
    fid = {k: i for i, (k, _) in enumerate(fams)}
    gid = np.array([fid[k] for k, v in fams for _ in v])
    pred = np.full(len(y), np.nan)
    for tr, te in LeaveOneGroupOut().split(X, y, gid):
        m = HistGradientBoostingRegressor(max_iter=300, max_depth=3, learning_rate=0.05,
                                          l2_regularization=3.0, min_samples_leaf=8, random_state=0).fit(X[tr], y[tr])
        pred[te] = m.predict(X[te])
    taus, ctaus = [], []
    for k, v in fams:
        mask = gid == fid[k]
        if mask.sum() >= 4 and np.std(y[mask]) > 0:
            t = spearmanr(pred[mask], y[mask]).statistic
            if not np.isnan(t):
                taus.append(t)
                if np.mean([abs(x["net_charge"]) for x in v]) >= 2:
                    ctaus.append(t)
    print(f"=== (A) SEQUENCE selectivity AT SCALE ({len(fams)} families, {len(y)} peptides) ===")
    print(f"  within-family tau = {np.mean(taus):+.3f}  (median {np.median(taus):+.3f}, n_fam={len(taus)})")
    print(f"  charged-family tau = {np.mean(ctaus):+.3f}  (n={len(ctaus)})")
    print(f"  frac tau>0 = {np.mean([t>0 for t in taus]):.2f}")
    print(f"  (branch 80-family seq baseline was +0.059 — does 5.7x scale move it?)\n")

    # (B) PPI-clone on PPIKB structured subset (branch features) vs us
    pf = [json.loads(l) for l in open(ROOT / "data/ppikb_features.jsonl") if json.loads(l).get("desc3d")]
    base = [json.loads(l) for l in open(ROOT / "data/e180_protdcal3d.jsonl") if json.loads(l).get("desc")]
    pf_pdbs = {p["pdb"].lower() for p in pf}
    base = [b for b in base if b["pdb"].lower() not in pf_pdbs]
    from sklearn.feature_selection import SelectKBest, f_regression
    from sklearn.pipeline import Pipeline
    from sklearn.preprocessing import StandardScaler
    from sklearn.svm import SVR
    Xtr = np.nan_to_num([b["desc"] for b in base]); ytr = np.array([b["y"] for b in base])
    clone = Pipeline([("sc", StandardScaler()), ("sel", SelectKBest(f_regression, k=37)),
                      ("svr", SVR(kernel="rbf", C=4.0, gamma="scale"))]).fit(Xtr, ytr)
    Xpf = np.nan_to_num([p["desc3d"] for p in pf]); ypf = np.array([p["y"] for p in pf])
    rclone, mclone = met(clone.predict(Xpf), ypf)
    print(f"=== (B) PPI-clone (ProtDCal-3D, trained on 925) ON PPIKB-structured subset (n={len(pf)}) ===")
    print(f"  PPI-clone vs truth: r={rclone:+.3f}  MAE={mclone:.2f}")
    print(f"  (clone on T100 was r~0.32 — if PPIKB is LOWER, PPIKB is a harder/noisier set for EVERYONE,")
    print(f"   which is why both we and the clone 'get worse' on it)")


if __name__ == "__main__":
    main()
